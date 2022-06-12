from django.shortcuts import render

from .models import Student, Xy
from .serializers import StudentSerializer, XySerializer, serializers
from rest_framework.generics import ListAPIView

from django.http import HttpResponse
import pandas as pd
import json
import openpyxl

# Create your views here.
class StudentList(ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

class XyList(ListAPIView):
    queryset = Xy.objects.all()
    serializer_class = XySerializer

def func1(request):
    Xy.objects.all().delete()   # Use this to delete all the values in the model
    
  
    # load excel with its path
    wrkbk = openpyxl.load_workbook("book3.xlsx")
  
    sh = wrkbk.active
    for i in range(1, sh.max_row+1):
        print(i)

        # data = pd.read_excel("book3.xlsx", skiprows=i, usecols="A", nrows=0)
        # x_val = str(data.columns[0])

        # data = pd.read_excel("book3.xlsx", skiprows=i, usecols="B", nrows=0)
        # y_val = str(data.columns[0])
        
        # xy_data = Xy(x=x_val, y=y_val)  
        # xy_data.save() 
        
        for j in range(1, sh.max_column+1):
            cell_obj = sh.cell(row=i, column=j)
            if(j==2):
                y_val = cell_obj.value; 
            elif(j==1):
                x_val = cell_obj.value

        xy_data = Xy(x=x_val, y=y_val)  
        xy_data.save() 
            # print(cell_obj.value[0], end=" ")
    return HttpResponse("<h1>Function is Working</h1>")