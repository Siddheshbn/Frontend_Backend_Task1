import pandas as pd
import openpyxl

wrkbk = openpyxl.load_workbook("book3.xlsx")

sh = wrkbk.active
for i in range(1, sh.max_row + 1):
    print(i)

    # data = pd.read_excel("book3.xlsx", skiprows=i, usecols="A", nrows=0)
    # x_val = str(data.columns[0])

    # data = pd.read_excel("book3.xlsx", skiprows=i, usecols="B", nrows=0)
    # y_val = str(data.columns[0])

    # xy_data = Xy(x=x_val, y=y_val)
    # xy_data.save()

    print("\n")
    print("Row ", i, " data :")

    for j in range(1, sh.max_column + 1):
        cell_obj = sh.cell(row=i, column=j)
        print(cell_obj.value, end=" ")
        print(type(cell_obj.value))